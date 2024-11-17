import pandas as pd
from typing import List, Iterable
from openpyxl import load_workbook, Workbook
from config import VOCABULARY as DATA_PATH
from config import VOCABULARY_STRUCTURE

""" LOADING SECTION """
_workbook: Workbook = None
sheet_names: tuple = None
__data_cache: dict = None

# TODO Order words by stats


def _reload_data():
    global _workbook, __data_cache, sheet_names

    print("Reloading the data")
    _workbook = load_workbook(DATA_PATH)
    sheet_names = _workbook.sheetnames
    __data_cache = dict()


if _workbook is None or __data_cache is None or sheet_names is None:
    from WordRecorder import load_stats
    _reload_data()
    load_stats()

""" LOADING SECTION """


def _row_to_list(row) -> List[str]:
    from openpyxl.cell import Cell

    return list(map(Cell.value.fget, row))


def _df_to_iter(df: pd.DataFrame) -> Iterable:
    df.reset_index()
    df = df.sample(frac=1).iterrows()
    return df


def _sheet_to_table(sheet_name: str, columns_cutoff: int) -> pd.DataFrame:
    if sheet_name in __data_cache:
        return __data_cache[sheet_name]

    sheet = _workbook[sheet_name]
    rows = iter(sheet.rows)
    columns = _row_to_list(next(rows))[:columns_cutoff]

    print(f"Loaded sheet: {sheet.title} (columns: {columns})")

    rows_list = []
    for row in rows:
        data = _row_to_list(row[:columns_cutoff])
        if not any(data):
            break
        rows_list.append(data)

    df = pd.DataFrame(rows_list, columns=list(map(str.lower, columns)), dtype=str)
    __data_cache[sheet_name] = df

    return df


def get_nouns() -> Iterable:
    df = _sheet_to_table(_workbook.sheetnames[VOCABULARY_STRUCTURE["noun"][0]], VOCABULARY_STRUCTURE["noun"][1])
    return _df_to_iter(df)


def get_verbs() -> Iterable:
    verbs = _sheet_to_table(_workbook.sheetnames[VOCABULARY_STRUCTURE["verb"][0]], VOCABULARY_STRUCTURE["verb"][1])
    return _df_to_iter(verbs)


def get_adjectives() -> Iterable:
    df = _sheet_to_table(_workbook.sheetnames[VOCABULARY_STRUCTURE["adjektiv"][0]], VOCABULARY_STRUCTURE['adjektiv'][1])
    return _df_to_iter(df)


def get_adverbs() -> Iterable:
    df = _sheet_to_table(_workbook.sheetnames[VOCABULARY_STRUCTURE["adverb"][0]], VOCABULARY_STRUCTURE['adverb'][1])
    return _df_to_iter(df)


def get_phrases():
    df = _sheet_to_table(_workbook.sheetnames[VOCABULARY_STRUCTURE["phrase"][0]], VOCABULARY_STRUCTURE['phrase'][1])
    return _df_to_iter(df)
